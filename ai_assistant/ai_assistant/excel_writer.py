from __future__ import annotations

from io import BytesIO
from datetime import date, datetime

from .account_mapping import account_root_type, base_account


BALANCE_ACCOUNTS = [
    "银行存款", "应收账款", "预付账款", "其他应收款", "库存商品", "固定资产", "累计折旧",
    "应付账款", "预收账款", "应付职工薪酬", "应交税费", "其他应付款", "实收资本", "利润分配",
    "主营业务收入", "其他业务收入", "主营业务成本", "其他业务成本", "税金及附加", "销售费用",
    "管理费用", "财务费用", "营业外支出", "所得税", "以前年度损益调整",
]


def _amount(value):
    value = float(value or 0)
    return round(value, 2) if abs(value) >= 0.005 else None


def _month_serial(transactions):
    for row in transactions:
        posting_date = row.get("posting_date")
        if isinstance(posting_date, (date, datetime)):
            return posting_date.strftime("%Y%m")
        text = str(posting_date or "")
        digits = "".join(ch for ch in text if ch.isdigit())
        if len(digits) >= 6:
            return digits[:6]
    return datetime.now().strftime("%Y%m")


def _style_sheet(ws, header_rows=(1,), freeze="A2"):
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    thin = Side(style="thin", color="D9E2EC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="EAF2F8")

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = border
            if cell.row in header_rows:
                cell.font = Font(bold=True)
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if isinstance(cell.value, (int, float)) and cell.column > 1:
                cell.number_format = '#,##0.00'

    ws.freeze_panes = freeze


def _write_vouchers(wb, voucher_rows):
    ws = wb.active
    ws.title = "凭证模板"
    headers = ["凭证号", "凭证组号", "分录行号", "摘要", "借方科目", "贷方科目", "借方金额", "贷方金额", "匹配来源", "复核状态", "AI置信度", "匹配说明"]
    ws.append(headers)
    for row in voucher_rows:
        ws.append([
            row.get("voucher_no"), row.get("group_id"), row.get("line_no"), row.get("summary"),
            row.get("debit_account"), row.get("credit_account"), _amount(row.get("debit_amount")),
            _amount(row.get("credit_amount")), row.get("mapping_source"), row.get("review_status"),
            row.get("ai_confidence"), row.get("mapping_reason"),
        ])
    widths = [10, 12, 10, 42, 24, 28, 14, 14, 12, 12, 12, 48]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width
    _style_sheet(ws)


def _write_trial_balance(wb, trial_balance):
    ws = wb.create_sheet("科目余额表")
    headers = ["科目名称", "期初借方", "期初贷方", "本期借方", "本期贷方", "累计借方", "累计贷方", "期末借方", "期末贷方"]
    ws.append(headers)

    written = set()
    ordered_accounts = BALANCE_ACCOUNTS + [acct for acct in sorted(trial_balance) if base_account(acct) not in BALANCE_ACCOUNTS and acct not in BALANCE_ACCOUNTS]
    for account in ordered_accounts:
        if account in written:
            continue
        data = trial_balance.get(account, {})
        if account not in trial_balance and "-" in account:
            continue
        ws.append([
            account,
            _amount(data.get("opening_debit")), _amount(data.get("opening_credit")),
            _amount(data.get("period_debit")), _amount(data.get("period_credit")),
            _amount(data.get("cumulative_debit")), _amount(data.get("cumulative_credit")),
            _amount(data.get("ending_debit")), _amount(data.get("ending_credit")),
        ])
        written.add(account)

    widths = [28, 14, 14, 14, 14, 14, 14, 14, 14]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width
    _style_sheet(ws)


def _ending_debit(trial_balance, account):
    return float(trial_balance.get(account, {}).get("ending_debit") or 0)


def _ending_credit(trial_balance, account):
    return float(trial_balance.get(account, {}).get("ending_credit") or 0)


def _period_debit(trial_balance, account):
    return float(trial_balance.get(account, {}).get("period_debit") or 0)


def _period_credit(trial_balance, account):
    return float(trial_balance.get(account, {}).get("period_credit") or 0)


def _opening_credit(trial_balance, account):
    return float(trial_balance.get(account, {}).get("opening_credit") or 0)


def _opening_debit(trial_balance, account):
    return float(trial_balance.get(account, {}).get("opening_debit") or 0)


def _net_asset_balance(trial_balance, account):
    return _ending_debit(trial_balance, account) - _ending_credit(trial_balance, account)


def _net_liability_balance(trial_balance, account):
    return _ending_credit(trial_balance, account) - _ending_debit(trial_balance, account)


def _net_income_balance(trial_balance, account):
    return _period_credit(trial_balance, account) - _period_debit(trial_balance, account)


def _net_expense_balance(trial_balance, account):
    return _period_debit(trial_balance, account) - _period_credit(trial_balance, account)


def _net_profit_amount(trial_balance):
    revenue = _net_income_balance(trial_balance, "主营业务收入") + _net_income_balance(trial_balance, "其他业务收入")
    cost = _net_expense_balance(trial_balance, "主营业务成本") + _net_expense_balance(trial_balance, "其他业务成本")
    tax_and_surcharges = _net_expense_balance(trial_balance, "税金及附加")
    selling = _net_expense_balance(trial_balance, "销售费用")
    admin = _net_expense_balance(trial_balance, "管理费用")
    finance = _net_expense_balance(trial_balance, "财务费用")
    non_operating_income = _net_income_balance(trial_balance, "营业外收入")
    non_operating_expense = _net_expense_balance(trial_balance, "营业外支出")
    income_tax = _net_expense_balance(trial_balance, "所得税")
    return revenue - cost - tax_and_surcharges - selling - admin - finance + non_operating_income - non_operating_expense - income_tax


def _sum_accounts(trial_balance, accounts, side="debit"):
    if side == "credit":
        return sum(_ending_credit(trial_balance, account) for account in accounts)
    return sum(_ending_debit(trial_balance, account) for account in accounts)


def _root_total(trial_balance, root_type):
    total = 0.0
    for account in trial_balance:
        if account_root_type(account) != root_type:
            continue
        if root_type == "asset":
            total += _net_asset_balance(trial_balance, account)
        elif root_type in ("liability", "equity"):
            total += _net_liability_balance(trial_balance, account)
        elif root_type == "asset_credit":
            total -= _ending_credit(trial_balance, account) - _ending_debit(trial_balance, account)
    return total


def _write_balance_sheet(wb, trial_balance, company_name=None, bank_context=None):
    ws = wb.create_sheet("资产负债表")
    company = company_name or ""
    ws.append(["资产负债表", None, None, None, None, None, None, None])
    ws.append([None, None, None, None, None, None, None, "会企01表"])
    ws.append([f"编制单位：{company}", None, None, None, datetime.now().strftime("%Y-%m-%d"), None, None, "单位：元"])
    ws.append(["资产", "行次", "期末余额", "年初余额", "负债和所有者权益", "行次", "期末余额", "年初余额"])

    bank_context = bank_context or {}
    monetary_funds = _net_asset_balance(trial_balance, "银行存款")
    receivables = _net_asset_balance(trial_balance, "应收账款")
    prepayments = _net_asset_balance(trial_balance, "预付账款")
    other_receivables = _net_asset_balance(trial_balance, "其他应收款")
    inventory = _net_asset_balance(trial_balance, "库存商品")
    fixed_assets = _net_asset_balance(trial_balance, "固定资产") - _ending_credit(trial_balance, "累计折旧")
    current_assets = monetary_funds + receivables + prepayments + other_receivables + inventory
    total_assets = _root_total(trial_balance, "asset") + _root_total(trial_balance, "asset_credit")

    payables = _net_liability_balance(trial_balance, "应付账款")
    advances = _net_liability_balance(trial_balance, "预收账款")
    payroll = _net_liability_balance(trial_balance, "应付职工薪酬")
    taxes = _net_liability_balance(trial_balance, "应交税费")
    other_payables = _net_liability_balance(trial_balance, "其他应付款")
    current_liabilities = payables + advances + payroll + taxes + other_payables
    total_liabilities = _root_total(trial_balance, "liability")
    capital = _net_liability_balance(trial_balance, "实收资本")
    opening_retained = _opening_credit(trial_balance, "利润分配") - _opening_debit(trial_balance, "利润分配")
    current_year_profit = _net_liability_balance(trial_balance, "本年利润")
    net_profit = _net_profit_amount(trial_balance)
    retained = opening_retained + current_year_profit + net_profit
    total_equity = _root_total(trial_balance, "equity") + net_profit
    total_liabilities_equity = total_liabilities + total_equity

    rows = [
        ["流动资产：", None, None, None, "流动负债：", None, None, None],
        ["货币资金", 1, _amount(monetary_funds), None, "短期借款", 32, None, None],
        ["应收账款", 2, _amount(receivables), None, "应付账款", 33, _amount(payables), None],
        ["预付款项", 3, _amount(prepayments), None, "预收账款", 34, _amount(advances), None],
        ["其他应收款", 4, _amount(other_receivables), None, "应付职工薪酬", 35, _amount(payroll), None],
        ["存货", 5, _amount(inventory), None, "应交税费", 36, _amount(taxes), None],
        ["流动资产合计", 6, _amount(current_assets), None, "其他应付款", 37, _amount(other_payables), None],
        ["非流动资产：", None, None, None, "流动负债合计", 38, _amount(current_liabilities), None],
        ["固定资产净额", 7, _amount(fixed_assets), None, "非流动负债：", None, None, None],
        ["非流动资产合计", 8, _amount(fixed_assets), None, "非流动负债合计", 39, _amount(total_liabilities - current_liabilities), None],
        ["资产总计", 9, _amount(total_assets), None, "所有者权益：", None, None, None],
        [None, None, None, None, "实收资本", 40, _amount(capital), None],
        [None, None, None, None, "未分配利润/本年利润", 41, _amount(retained), None],
        [None, None, None, None, "所有者权益合计", 42, _amount(total_equity), None],
        [None, None, None, None, "负债和所有者权益总计", 43, _amount(total_liabilities_equity), None],
    ]
    for row in rows:
        ws.append(row)

    for col in "ABCDEFGH":
        ws.column_dimensions[col].width = 18
    ws.merge_cells("A1:H1")
    _style_sheet(ws, header_rows=(1, 4), freeze="A5")


def _write_income_statement(wb, trial_balance, company_name=None):
    ws = wb.create_sheet("利润表")
    company = company_name or ""
    revenue = _net_income_balance(trial_balance, "主营业务收入") + _net_income_balance(trial_balance, "其他业务收入")
    cost = _net_expense_balance(trial_balance, "主营业务成本") + _net_expense_balance(trial_balance, "其他业务成本")
    tax_and_surcharges = _net_expense_balance(trial_balance, "税金及附加")
    selling = _net_expense_balance(trial_balance, "销售费用")
    admin = _net_expense_balance(trial_balance, "管理费用")
    finance = _net_expense_balance(trial_balance, "财务费用")
    non_operating_income = _net_income_balance(trial_balance, "营业外收入")
    non_operating_expense = _net_expense_balance(trial_balance, "营业外支出")
    income_tax = _net_expense_balance(trial_balance, "所得税")
    operating_profit = revenue - cost - tax_and_surcharges - selling - admin - finance
    total_profit = operating_profit + non_operating_income - non_operating_expense
    net_profit = total_profit - income_tax

    ws.append(["利润表", None, None, None])
    ws.append(["", None, "", "会企02表"])
    ws.append([f"编制单位：{company}", None, datetime.now().strftime("%Y-%m-%d"), "单位：元"])
    ws.append(["项目", "行次", "本期金额", "上期金额"])
    rows = [
        ["一、营业收入", 1, _amount(revenue), None],
        ["减：营业成本", 2, _amount(cost), None],
        ["税金及附加", 3, _amount(tax_and_surcharges), None],
        ["销售费用", 4, _amount(selling), None],
        ["管理费用", 5, _amount(admin), None],
        ["财务费用", 6, _amount(finance), None],
        ["二、营业利润", 7, _amount(operating_profit), None],
        ["加：营业外收入", 8, _amount(non_operating_income), None],
        ["减：营业外支出", 9, _amount(non_operating_expense), None],
        ["三、利润总额", 10, _amount(total_profit), None],
        ["减：所得税费用", 11, _amount(income_tax), None],
        ["四、净利润", 12, _amount(net_profit), None],
    ]
    for row in rows:
        ws.append(row)
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.merge_cells("A1:D1")
    _style_sheet(ws, header_rows=(1, 4), freeze="A5")


def build_financial_report_workbook(voucher_rows, trial_balance, transactions, company_name=None, bank_context=None):
    from openpyxl import Workbook

    wb = Workbook()
    _write_vouchers(wb, voucher_rows)
    _write_trial_balance(wb, trial_balance)
    _write_balance_sheet(wb, trial_balance, company_name=company_name, bank_context=bank_context)
    _write_income_statement(wb, trial_balance, company_name=company_name)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue(), _month_serial(transactions)
