from __future__ import annotations

import os
import re
import uuid
from difflib import SequenceMatcher
from collections import defaultdict
from datetime import datetime
from decimal import Decimal, InvalidOperation

import frappe
from frappe.utils.file_manager import get_file_path, save_file

from .account_mapping import account_root_type, base_account, is_valid_account, map_transaction
from .excel_writer import build_financial_report_workbook

REQUIRED_HEADERS = ["入账日期", "转出金额", "转入金额", "余额", "对方单位", "对方账号", "摘要", "用途"]
CORE_HEADERS = {"入账日期", "转出金额", "转入金额", "余额", "对方单位", "摘要"}
HEADER_SYNONYMS = {
    "入账日期": [
        "入账日期", "交易日期", "记账日期", "记账时间", "交易时间", "日期", "发生日期", "账务日期",
        "交易日", "交易日期时间", "交易时间戳", "Date", "Value Date", "Transaction Date", "Posting Date",
    ],
    "转出金额": [
        "转出金额", "支出金额", "付款金额", "借方金额", "借方发生额", "支出", "借方", "转出", "付款",
        "扣款金额", "取出金额", "Debit", "Debit Amount", "Withdrawal", "Paid Amount",
    ],
    "转入金额": [
        "转入金额", "收入金额", "收款金额", "贷方金额", "贷方发生额", "收入", "贷方", "转入", "收款",
        "存入金额", "Credit", "Credit Amount", "Deposit", "Received Amount",
    ],
    "余额": [
        "余额", "账户余额", "可用余额", "当前余额", "期末余额", "本次余额", "Balance",
        "Available Balance", "Closing Balance",
    ],
    "对方单位": [
        "对方单位", "对方名称", "对方户名", "收款人", "付款人", "交易对手", "对手方", "对方账户名称",
        "对方账号名称", "对方户名/账户名称", "交易对方", "Counterparty", "Payee", "Beneficiary",
        "Payer", "Counterparty Name",
    ],
    "对方账号": [
        "对方账号", "对方账户", "对方帐号", "收款账号", "付款账号", "对手账号", "对方卡号",
        "Counterparty Account", "Payee Account", "Beneficiary Account",
    ],
    "摘要": [
        "摘要", "备注", "交易摘要", "交易描述", "用途摘要", "摘要信息", "交易备注", "交易附言",
        "Description", "Memo", "Narration", "Remark", "Remarks",
    ],
    "用途": [
        "用途", "交易用途", "用途说明", "付款用途", "业务用途", "Purpose", "Payment Purpose",
    ],
}


def _normalize_header(value):
    text = _clean(value).lower()
    text = re.sub(r"[\s　:：/\\()（）\[\]【】._\-]+", "", text)
    return text


def _header_match_score(cell, synonym):
    cell_norm = _normalize_header(cell)
    syn_norm = _normalize_header(synonym)
    if not cell_norm or not syn_norm:
        return 0.0
    if cell_norm == syn_norm:
        return 1.0
    if cell_norm in syn_norm or syn_norm in cell_norm:
        return 0.86
    return SequenceMatcher(None, cell_norm, syn_norm).ratio()


def _is_valid_header_map(matched):
    has_date = "入账日期" in matched
    has_amount = "转出金额" in matched or "转入金额" in matched
    support_hits = sum(1 for header in ("余额", "对方单位", "摘要") if header in matched)
    return has_date and has_amount and support_hits >= 2


def _match_headers(cell_values, threshold=0.6):
    result = {}
    used_cols = set()

    for canonical, synonyms in HEADER_SYNONYMS.items():
        for col_idx, cell in enumerate(cell_values):
            if col_idx in used_cols:
                continue
            if any(_normalize_header(cell) == _normalize_header(syn) for syn in synonyms):
                result[canonical] = col_idx + 1
                used_cols.add(col_idx)
                break

    for canonical in REQUIRED_HEADERS:
        if canonical in result:
            continue
        best_score = 0.0
        best_col = None
        for col_idx, cell in enumerate(cell_values):
            if col_idx in used_cols or not _clean(cell):
                continue
            score = max(_header_match_score(cell, syn) for syn in HEADER_SYNONYMS.get(canonical, [canonical]))
            if score > best_score:
                best_score = score
                best_col = col_idx
        if best_col is not None and best_score >= threshold:
            result[canonical] = best_col + 1
            used_cols.add(best_col)

    return result


def _validate_ai_header_map(raw_mapping, max_col):
    if not isinstance(raw_mapping, dict):
        return {}
    result = {}
    used_cols = set()
    for header in REQUIRED_HEADERS:
        if header not in raw_mapping:
            continue
        try:
            col_idx = int(raw_mapping[header]) + 1
        except (TypeError, ValueError):
            continue
        if col_idx < 1 or col_idx > max_col or col_idx in used_cols:
            continue
        result[header] = col_idx
        used_cols.add(col_idx)
    return result


def _to_amount(value):
    if value is None:
        return Decimal("0")
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value)).quantize(Decimal("0.01"))

    text = str(value).strip()
    if not text:
        return Decimal("0")

    negative = False
    if text.startswith("(") and text.endswith(")"):
        negative = True
        text = text[1:-1]
    if text.startswith("-"):
        negative = True
        text = text[1:]

    multiplier = Decimal("1")
    if "万元" in text:
        multiplier = Decimal("10000")
    text = text.replace("万元", "").replace("元", "")
    text = re.sub(r"[￥¥,，\s　]", "", text)
    if not text:
        return Decimal("0")

    try:
        amount = (Decimal(text) * multiplier).quantize(Decimal("0.01"))
        return -amount if negative else amount
    except InvalidOperation:
        return Decimal("0")


def _clean(value):
    if value is None:
        return ""
    return str(value).strip()


def _build_merged_lookup(ws):
    merged_ranges = list(getattr(ws.merged_cells, "ranges", []) or [])
    if not merged_ranges:
        return None
    merged_lookup = {}
    for merged_range in merged_ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        top_left = ws.cell(row=min_row, column=min_col).value
        for row_idx in range(min_row, max_row + 1):
            for col_idx in range(min_col, max_col + 1):
                merged_lookup[(row_idx, col_idx)] = top_left
    return merged_lookup


def _cell_value(ws, row_idx, col_idx, merged_lookup=None):
    value = ws.cell(row=row_idx, column=col_idx).value
    if value is not None:
        return value
    if merged_lookup:
        return merged_lookup.get((row_idx, col_idx))
    return None


def _has_valid_data_row(ws, header_row, columns, merged_lookup=None):
    max_check_row = min(ws.max_row or header_row, header_row + 10)
    for row_idx in range(header_row + 1, max_check_row + 1):
        row = {header: _cell_value(ws, row_idx, col_idx, merged_lookup) for header, col_idx in columns.items()}
        out_amount = _to_amount(row.get("转出金额"))
        in_amount = _to_amount(row.get("转入金额"))
        if (out_amount > 0 or in_amount > 0) and _parse_posting_date(row.get("入账日期")) is not None:
            return True
    return False


def _find_header_row(ws, header_classifier=None, merged_lookup=None):
    max_rows = min(ws.max_row or 1, 50)
    max_cols = ws.max_column or 1
    merged_lookup = merged_lookup or {}
    scanned = []
    best_row = None
    best_columns = {}

    for row_idx in range(1, max_rows + 1):
        values = [_clean(_cell_value(ws, row_idx, col_idx, merged_lookup)) for col_idx in range(1, max_cols + 1)]
        if not any(values):
            continue
        matched = _match_headers(values)
        scanned.append({"row": row_idx, "values": values, "matched": matched})
        if len(matched) > len(best_columns):
            best_row, best_columns = row_idx, matched
        if _is_valid_header_map(matched) and _has_valid_data_row(ws, row_idx, matched, merged_lookup):
            return row_idx, matched

    if header_classifier:
        for item in scanned[:10]:
            try:
                ai_mapping = _validate_ai_header_map(header_classifier(item["values"]), max_cols)
            except Exception:
                ai_mapping = {}
            if _is_valid_header_map(ai_mapping) and _has_valid_data_row(ws, item["row"], ai_mapping, merged_lookup):
                return item["row"], ai_mapping

    preview = []
    for item in scanned[:8]:
        non_empty = [value for value in item["values"] if value][:12]
        matched_names = ", ".join(sorted(item["matched"])) or "无"
        preview.append(f"第 {item['row']} 行：{non_empty}；已识别：{matched_names}")
    detail = "\n".join(preview) or "前 50 行未发现非空候选表头。"
    best_names = ", ".join(sorted(best_columns)) or "无"
    raise ValueError(
        "未找到可用的银行交易明细表头。已尝试同义词匹配和模糊匹配，"
        f"最佳候选行为第 {best_row or '-'} 行，已识别字段：{best_names}。\n"
        "请确认文件至少包含日期列、收入或支出金额列，以及余额/对方单位/摘要中的任意两类信息。\n"
        f"扫描摘要：\n{detail}"
    )


def parse_bank_transactions(source_path, header_classifier=None):
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("当前运行环境缺少 openpyxl，无法读取 Excel。请在 bench 环境安装 openpyxl。") from exc

    wb = openpyxl.load_workbook(source_path, data_only=True, read_only=False)
    ws = wb.worksheets[0]
    merged_lookup = _build_merged_lookup(ws)
    header_row, columns = _find_header_row(ws, header_classifier=header_classifier, merged_lookup=merged_lookup)
    transactions = []

    for row_idx in range(header_row + 1, (ws.max_row or header_row) + 1):
        row = {header: _cell_value(ws, row_idx, col_idx, merged_lookup) for header, col_idx in columns.items()}
        out_amount = _to_amount(row.get("转出金额"))
        in_amount = _to_amount(row.get("转入金额"))
        posting_date = row.get("入账日期")
        if out_amount <= 0 and in_amount <= 0:
            continue
        if _parse_posting_date(posting_date) is None:
            continue

        transactions.append({
            "row_no": row_idx,
            "posting_date": posting_date,
            "out_amount": out_amount,
            "in_amount": in_amount,
            "balance": _to_amount(row.get("余额")),
            "counterparty": _clean(row.get("对方单位")),
            "counterparty_account": _clean(row.get("对方账号")),
            "summary": _clean(row.get("摘要")),
            "purpose": _clean(row.get("用途")),
        })

    if not transactions:
        raise ValueError("Excel 中没有识别到有效的转入/转出交易记录。")
    return transactions


def _summary_for_transaction(tx):
    summary = tx.get("summary") or tx.get("purpose") or "银行交易"
    counterparty = tx.get("counterparty")
    if counterparty and counterparty not in summary and len(summary) < 18:
        return f"{summary}-{counterparty[:18]}"
    return summary


def _build_mapped_entries(transactions):
    mapped_entries = []
    review_notes = []
    candidates = []

    for idx, tx in enumerate(transactions, start=1):
        out_amount = tx["out_amount"]
        in_amount = tx["in_amount"]
        if out_amount > 0 and in_amount > 0:
            review_notes.append(f"第 {tx['row_no']} 行同时存在转入和转出金额，已按两笔方向分别处理。")

        directions = []
        if out_amount > 0:
            directions.append(("out", out_amount))
        if in_amount > 0:
            directions.append(("in", in_amount))

        for direction, amount in directions:
            debit_account, credit_account, reason = map_transaction(
                tx.get("summary"), tx.get("purpose"), tx.get("counterparty"), direction=direction
            )
            candidate_id = f"{idx}-{direction}-{tx['row_no']}"
            entry = {
                "candidate_id": candidate_id,
                "voucher_no": idx,
                "tx": tx,
                "direction": direction,
                "amount": amount,
                "debit_account": debit_account,
                "credit_account": credit_account,
                "mapping_reason": reason,
                "needs_review": "需人工复核" in reason,
                "ai_applied": False,
                "ai_confidence": None,
            }
            mapped_entries.append(entry)

            if entry["needs_review"]:
                candidates.append({
                    "id": candidate_id,
                    "source_row": tx["row_no"],
                    "direction": direction,
                    "amount": float(amount),
                    "posting_date": str(tx.get("posting_date") or ""),
                    "summary": tx.get("summary") or "",
                    "purpose": tx.get("purpose") or "",
                    "counterparty": tx.get("counterparty") or "",
                    "counterparty_account": tx.get("counterparty_account") or "",
                    "fallback_debit_account": debit_account,
                    "fallback_credit_account": credit_account,
                })

    return mapped_entries, candidates, review_notes


def _valid_ai_suggestion(direction, debit_account, credit_account, amount=None, summary=None):
    if not is_valid_account(debit_account) or not is_valid_account(credit_account):
        return False
    if direction == "out" and credit_account != "银行存款":
        return False
    if direction == "in" and debit_account != "银行存款":
        return False
    return True


def _ai_suggestion_warnings(entry, debit_account, credit_account):
    warnings = []
    amount = Decimal(str(entry.get("amount") or 0))
    summary_text = _clean(entry["tx"].get("summary")) + _clean(entry["tx"].get("purpose"))
    if amount > Decimal("10000000"):
        warnings.append(f"单笔金额超过 1000 万：{amount:.2f}")
    if re.search(r"工资|薪酬|薪资", summary_text) and base_account(debit_account) != "应付职工薪酬":
        warnings.append("摘要疑似工资薪酬，但 AI 返回科目不是应付职工薪酬")
    return warnings


def _apply_ai_suggestions(mapped_entries, ai_classifier=None):
    candidates = [
        {
            "id": entry["candidate_id"],
            "source_row": entry["tx"]["row_no"],
            "direction": entry["direction"],
            "amount": float(entry["amount"]),
            "posting_date": str(entry["tx"].get("posting_date") or ""),
            "summary": entry["tx"].get("summary") or "",
            "purpose": entry["tx"].get("purpose") or "",
            "counterparty": entry["tx"].get("counterparty") or "",
            "counterparty_account": entry["tx"].get("counterparty_account") or "",
            "fallback_debit_account": entry["debit_account"],
            "fallback_credit_account": entry["credit_account"],
        }
        for entry in mapped_entries
        if entry["needs_review"]
    ]
    if not candidates or not ai_classifier:
        return {"candidate_count": len(candidates), "applied_count": 0, "invalid_count": 0, "error": None}

    try:
        suggestions = ai_classifier(candidates) or {}
    except Exception as exc:
        return {"candidate_count": len(candidates), "applied_count": 0, "invalid_count": 0, "error": str(exc)}

    applied_count = 0
    invalid_count = 0
    for entry in mapped_entries:
        suggestion = suggestions.get(entry["candidate_id"])
        if not suggestion:
            continue

        debit_account = _clean(suggestion.get("debit_account"))
        credit_account = _clean(suggestion.get("credit_account"))
        if not _valid_ai_suggestion(entry["direction"], debit_account, credit_account, entry["amount"], entry["tx"].get("summary")):
            invalid_count += 1
            continue

        entry["debit_account"] = debit_account
        entry["credit_account"] = credit_account
        confidence = float(suggestion.get("confidence") or 0)
        warnings = _ai_suggestion_warnings(entry, debit_account, credit_account)
        warning_text = f"；{'；'.join(warnings)}" if warnings else ""
        entry["mapping_reason"] = f"AI辅助判断：{_clean(suggestion.get('reason')) or '未提供原因'}{warning_text}"
        entry["needs_review"] = confidence < 0.75 or bool(warnings)
        entry["ai_applied"] = True
        entry["ai_confidence"] = confidence
        applied_count += 1

    return {"candidate_count": len(candidates), "applied_count": applied_count, "invalid_count": invalid_count, "error": None}


def validate_voucher_entry_balance(mapped_entries):
    notes = []
    for entry in mapped_entries:
        amount = Decimal(str(entry.get("amount") or 0)).quantize(Decimal("0.01"))
        if amount <= 0:
            entry["needs_review"] = True
            notes.append(f"凭证 {entry['voucher_no']}（源文件第 {entry['tx']['row_no']} 行）金额无效，需人工复核。")
        if not entry.get("debit_account") or not entry.get("credit_account"):
            entry["needs_review"] = True
            notes.append(f"凭证 {entry['voucher_no']}（源文件第 {entry['tx']['row_no']} 行）借贷科目不完整，需人工复核。")
    return notes




def _append_voucher_line(voucher_rows, entry, summary, debit_account=None, credit_account=None, debit_amount=None, credit_amount=None, mapping_reason=None, is_closing=False):
    voucher_rows.append({
        "voucher_no": entry["voucher_no"] if isinstance(entry, dict) else entry,
        "summary": summary,
        "debit_account": debit_account,
        "credit_account": credit_account,
        "debit_amount": float(debit_amount or 0) if debit_amount is not None else None,
        "credit_amount": float(credit_amount or 0) if credit_amount is not None else None,
        "source_row": entry.get("tx", {}).get("row_no") if isinstance(entry, dict) else None,
        "mapping_reason": mapping_reason or (entry.get("mapping_reason") if isinstance(entry, dict) else ""),
        "needs_review": bool(entry.get("needs_review")) if isinstance(entry, dict) else False,
        "ai_applied": bool(entry.get("ai_applied")) if isinstance(entry, dict) else False,
        "ai_confidence": entry.get("ai_confidence") if isinstance(entry, dict) else None,
        "mapping_source": "AI" if isinstance(entry, dict) and entry.get("ai_applied") else "规则",
        "review_status": "需复核" if isinstance(entry, dict) and entry.get("needs_review") else "已匹配",
        "is_closing": is_closing,
    })


def _payroll_benefit_accrual_account(account):
    if account == "应付职工薪酬-社保":
        return "管理费用-社保"
    if account == "应付职工薪酬-公积金":
        return "管理费用-公积金"
    return None


def _append_payroll_benefit_accrual(voucher_rows, entry, amount_float):
    expense_account = _payroll_benefit_accrual_account(entry.get("debit_account"))
    if not expense_account or entry.get("direction") != "out":
        return False
    summary = f"计提{entry['debit_account'].split('-', 1)[1]}"
    reason = f"支付{entry['debit_account'].split('-', 1)[1]}时同步计提：借记{expense_account}，贷记{entry['debit_account']}"
    _append_voucher_line(voucher_rows, entry, summary, debit_account=expense_account, debit_amount=amount_float, mapping_reason=reason)
    _append_voucher_line(voucher_rows, entry, None, credit_account=entry["debit_account"], credit_amount=amount_float, mapping_reason=reason)
    return True


def _append_profit_loss_closing_entries(voucher_rows):
    account_totals = defaultdict(lambda: {"debit": 0.0, "credit": 0.0})
    for row in voucher_rows:
        if row.get("is_closing"):
            continue
        debit_account = row.get("debit_account")
        credit_account = row.get("credit_account")
        if debit_account:
            base = base_account(debit_account)
            if account_root_type(base) in ("income", "expense"):
                account_totals[base]["debit"] += float(row.get("debit_amount") or 0)
        if credit_account:
            base = base_account(credit_account)
            if account_root_type(base) in ("income", "expense"):
                account_totals[base]["credit"] += float(row.get("credit_amount") or 0)

    next_voucher_no = max([int(row.get("voucher_no") or 0) for row in voucher_rows] or [0]) + 1
    closing_count = 0
    for account in sorted(account_totals):
        root = account_root_type(account)
        debit = account_totals[account]["debit"]
        credit = account_totals[account]["credit"]
        if root == "income":
            amount = round(credit - debit, 2)
            if amount <= 0:
                continue
            entry = {"voucher_no": next_voucher_no, "needs_review": False, "ai_applied": False, "ai_confidence": None, "tx": {}}
            reason = f"月末损益结转：{account} 转入 本年利润"
            _append_voucher_line(voucher_rows, entry, f"结转本期损益-{account}", debit_account=account, debit_amount=amount, mapping_reason=reason, is_closing=True)
            _append_voucher_line(voucher_rows, entry, None, credit_account="本年利润", credit_amount=amount, mapping_reason=reason, is_closing=True)
        elif root == "expense":
            amount = round(debit - credit, 2)
            if amount <= 0:
                continue
            entry = {"voucher_no": next_voucher_no, "needs_review": False, "ai_applied": False, "ai_confidence": None, "tx": {}}
            reason = f"月末损益结转：{account} 转入 本年利润"
            _append_voucher_line(voucher_rows, entry, f"结转本期损益-{account}", debit_account="本年利润", debit_amount=amount, mapping_reason=reason, is_closing=True)
            _append_voucher_line(voucher_rows, entry, None, credit_account=account, credit_amount=amount, mapping_reason=reason, is_closing=True)
        else:
            continue
        next_voucher_no += 1
        closing_count += 1
    return closing_count

def build_voucher_rows(transactions, ai_classifier=None):
    voucher_rows = []
    mapped_entries, _candidates, review_notes = _build_mapped_entries(transactions)
    ai_stats = _apply_ai_suggestions(mapped_entries, ai_classifier=ai_classifier)
    review_notes.extend(validate_voucher_entry_balance(mapped_entries))

    for entry in mapped_entries:
        tx = entry["tx"]
        amount_float = float(entry["amount"])
        summary = _summary_for_transaction(tx)
        if entry["needs_review"]:
            review_notes.append(f"凭证 {entry['voucher_no']}（源文件第 {tx['row_no']} 行）需人工复核：{summary}；{entry['mapping_reason']}")

        _append_voucher_line(
            voucher_rows, entry, summary, debit_account=entry["debit_account"],
            debit_amount=amount_float, mapping_reason=entry["mapping_reason"]
        )
        _append_voucher_line(
            voucher_rows, entry, None, credit_account=entry["credit_account"],
            credit_amount=amount_float, mapping_reason=entry["mapping_reason"]
        )
        _append_payroll_benefit_accrual(voucher_rows, entry, amount_float)

    closing_count = _append_profit_loss_closing_entries(voucher_rows)
    if closing_count:
        review_notes.append(f"已自动生成 {closing_count} 组月末损益结转分录，结转至本年利润。")

    if ai_stats["error"]:
        review_notes.append(f"AI辅助判断失败，已使用规则兜底：{ai_stats['error']}")
    if ai_stats["invalid_count"]:
        review_notes.append(f"AI返回 {ai_stats['invalid_count']} 条不合法科目或方向，已丢弃并使用规则兜底。")

    return voucher_rows, review_notes, ai_stats


def _parse_posting_date(value):
    if isinstance(value, datetime):
        return value.date()
    text = str(value or "").strip()
    if not text:
        return None
    normalized = re.sub(r"年|月", "-", text.replace("日", ""))
    normalized = re.sub(r"\s+", " ", normalized).strip()
    for candidate in (text, normalized):
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y%m%d", "%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S"):
            try:
                return datetime.strptime(candidate, fmt).date()
            except ValueError:
                pass
    return None


def _chronological_transactions(transactions):
    return sorted(
        transactions,
        key=lambda tx: (_parse_posting_date(tx.get("posting_date")) or datetime.min.date(), -(tx.get("row_no") or 0)),
    )


def validate_bank_balance_sequence(transactions):
    if len(transactions) < 2:
        return []
    notes = []
    chronological = _chronological_transactions(transactions)
    previous = chronological[0]
    previous_balance = previous.get("balance") or Decimal("0")
    for tx in chronological[1:]:
        expected = previous_balance + tx.get("in_amount", Decimal("0")) - tx.get("out_amount", Decimal("0"))
        actual = tx.get("balance") or Decimal("0")
        if abs(expected - actual) > Decimal("0.01"):
            notes.append(
                f"源文件第 {tx.get('row_no')} 行余额连续性异常：按上一笔余额推算应为 {expected:.2f}，实际为 {actual:.2f}，请核对是否缺行或排序异常。"
            )
        previous_balance = actual
    return notes


def build_bank_balance_context(transactions):
    if not transactions:
        return {"opening_balance": 0.0, "ending_balance": 0.0, "has_bank_balance": False}

    # Bank statements are usually exported newest-first. For same-day rows,
    # larger source row numbers happened earlier in chronological order.
    chronological = _chronological_transactions(transactions)
    first_tx = chronological[0]
    last_tx = chronological[-1]
    first_balance_after = first_tx.get("balance") or Decimal("0")
    opening_balance = first_balance_after - first_tx.get("in_amount", Decimal("0")) + first_tx.get("out_amount", Decimal("0"))
    ending_balance = last_tx.get("balance") or Decimal("0")

    return {
        "opening_balance": float(opening_balance),
        "ending_balance": float(ending_balance),
        "first_row_no": first_tx.get("row_no"),
        "last_row_no": last_tx.get("row_no"),
        "has_bank_balance": True,
    }


def build_trial_balance(voucher_rows, bank_context=None):
    totals = defaultdict(lambda: {
        "opening_debit": 0.0, "opening_credit": 0.0,
        "period_debit": 0.0, "period_credit": 0.0,
        "statement_debit": 0.0, "statement_credit": 0.0,
        "closing_debit": 0.0, "closing_credit": 0.0,
    })

    bank_context = bank_context or {}
    opening_bank = float(bank_context.get("opening_balance") or 0)
    if bank_context.get("has_bank_balance") and opening_bank:
        if opening_bank > 0:
            totals["银行存款"]["opening_debit"] += opening_bank
            totals["利润分配"]["opening_credit"] += opening_bank
        else:
            totals["银行存款"]["opening_credit"] += abs(opening_bank)
            totals["利润分配"]["opening_debit"] += abs(opening_bank)

    for row in voucher_rows:
        debit_account = row.get("debit_account")
        credit_account = row.get("credit_account")
        is_closing = bool(row.get("is_closing"))
        if debit_account:
            base = base_account(debit_account)
            amount = float(row.get("debit_amount") or 0)
            totals[base]["period_debit"] += amount
            if is_closing:
                totals[base]["closing_debit"] += amount
            else:
                totals[base]["statement_debit"] += amount
        if credit_account:
            base = base_account(credit_account)
            amount = float(row.get("credit_amount") or 0)
            totals[base]["period_credit"] += amount
            if is_closing:
                totals[base]["closing_credit"] += amount
            else:
                totals[base]["statement_credit"] += amount

    for account, data in totals.items():
        cumulative_debit = data["opening_debit"] + data["period_debit"]
        cumulative_credit = data["opening_credit"] + data["period_credit"]
        data["cumulative_debit"] = cumulative_debit
        data["cumulative_credit"] = cumulative_credit
        net = cumulative_debit - cumulative_credit
        data["ending_debit"] = net if net > 0 else 0.0
        data["ending_credit"] = -net if net < 0 else 0.0

    return dict(totals)


def _tb_ending_debit(trial_balance, account):
    return float(trial_balance.get(account, {}).get("ending_debit") or 0)


def _tb_ending_credit(trial_balance, account):
    return float(trial_balance.get(account, {}).get("ending_credit") or 0)


def _tb_period_debit(trial_balance, account):
    data = trial_balance.get(account, {})
    return float(data.get("statement_debit") if data.get("statement_debit") is not None else data.get("period_debit") or 0)


def _tb_period_credit(trial_balance, account):
    data = trial_balance.get(account, {})
    return float(data.get("statement_credit") if data.get("statement_credit") is not None else data.get("period_credit") or 0)


def _tb_net_asset(trial_balance, account):
    return _tb_ending_debit(trial_balance, account) - _tb_ending_credit(trial_balance, account)


def _tb_net_liability(trial_balance, account):
    return _tb_ending_credit(trial_balance, account) - _tb_ending_debit(trial_balance, account)


def _tb_root_total(trial_balance, root_type):
    total = 0.0
    for account in trial_balance:
        root = account_root_type(account)
        if root != root_type:
            continue
        if root_type == "asset":
            total += _tb_net_asset(trial_balance, account)
        elif root_type in ("liability", "equity"):
            total += _tb_net_liability(trial_balance, account)
        elif root_type == "asset_credit":
            total -= _tb_ending_credit(trial_balance, account) - _tb_ending_debit(trial_balance, account)
    return total


def _tb_net_profit(trial_balance):
    revenue = (_tb_period_credit(trial_balance, "主营业务收入") - _tb_period_debit(trial_balance, "主营业务收入")) + (_tb_period_credit(trial_balance, "其他业务收入") - _tb_period_debit(trial_balance, "其他业务收入"))
    cost = (_tb_period_debit(trial_balance, "主营业务成本") - _tb_period_credit(trial_balance, "主营业务成本")) + (_tb_period_debit(trial_balance, "其他业务成本") - _tb_period_credit(trial_balance, "其他业务成本"))
    expenses = sum(_tb_period_debit(trial_balance, account) - _tb_period_credit(trial_balance, account) for account in ("税金及附加", "销售费用", "管理费用", "财务费用", "营业外支出", "所得税"))
    non_operating_income = _tb_period_credit(trial_balance, "营业外收入") - _tb_period_debit(trial_balance, "营业外收入")
    return revenue - cost - expenses + non_operating_income


def validate_balance_sheet_totals(trial_balance):
    total_assets = _tb_root_total(trial_balance, "asset") + _tb_root_total(trial_balance, "asset_credit")
    total_liabilities = _tb_root_total(trial_balance, "liability")
    total_equity = _tb_root_total(trial_balance, "equity")
    diff = round(total_assets - total_liabilities - total_equity, 2)
    if abs(diff) <= 0.01:
        return []
    return [f"资产负债表勾稽异常：资产总计 {total_assets:.2f}，负债和所有者权益总计 {(total_liabilities + total_equity):.2f}，差异 {diff:.2f}，请复核科目映射和期初余额。"]


def _resolve_source_path(file_url_or_name):
    file_url_or_name = _clean(file_url_or_name)
    if not file_url_or_name:
        raise ValueError("缺少上传文件路径。")
    if file_url_or_name.startswith(("/files/", "/private/files/")) or "/" not in file_url_or_name:
        path = get_file_path(file_url_or_name)
    else:
        path = file_url_or_name
    if not path or not os.path.exists(path):
        raise FileNotFoundError(f"找不到上传文件：{file_url_or_name}")
    return path


def _company_name():
    try:
        company = frappe.defaults.get_user_default("Company") or frappe.db.get_single_value("Global Defaults", "default_company")
        return company or ""
    except Exception:
        return ""


def generate_financial_vouchers(file_url_or_name, ai_classifier=None, header_classifier=None):
    source_path = _resolve_source_path(file_url_or_name)
    transactions = parse_bank_transactions(source_path, header_classifier=header_classifier)
    voucher_rows, review_notes, ai_stats = build_voucher_rows(transactions, ai_classifier=ai_classifier)
    review_notes.extend(validate_bank_balance_sequence(transactions))
    bank_context = build_bank_balance_context(transactions)
    trial_balance = build_trial_balance(voucher_rows, bank_context=bank_context)
    review_notes.extend(validate_balance_sheet_totals(trial_balance))
    workbook_bytes, month_serial = build_financial_report_workbook(
        voucher_rows, trial_balance, transactions, company_name=_company_name(), bank_context=bank_context
    )

    source_stem = os.path.splitext(os.path.basename(source_path))[0]
    safe_stem = re.sub(r"[^0-9A-Za-z\u4e00-\u9fff._-]+", "_", source_stem).strip("_") or "银行交易明细"
    output_name = f"{safe_stem}_财务凭证报表_{month_serial}_{uuid.uuid4().hex[:8]}.xlsx"
    file_doc = save_file(output_name, workbook_bytes, None, None, is_private=1)

    debit_total = sum(float(row.get("debit_amount") or 0) for row in voucher_rows)
    credit_total = sum(float(row.get("credit_amount") or 0) for row in voucher_rows)
    if round(debit_total - credit_total, 2) != 0:
        review_notes.append(f"借贷合计不平衡：借方 {debit_total:.2f}，贷方 {credit_total:.2f}")

    text = (
        f"已根据上传的银行交易明细生成财务凭证报表。\n\n"
        f"- 识别交易：{len(transactions)} 笔\n"
        f"- 生成凭证分录行：{len(voucher_rows)} 行\n"
        f"- 规则未命中候选：{ai_stats['candidate_count']} 条\n"
        f"- AI辅助采纳：{ai_stats['applied_count']} 条\n"
        f"- 借方合计：￥{debit_total:,.2f}\n"
        f"- 贷方合计：￥{credit_total:,.2f}\n"
        f"- 输出文件：{file_doc.file_name}\n"
    )
    if review_notes:
        preview = "\n".join(f"  - {note}" for note in review_notes[:10])
        more = f"\n  - 另有 {len(review_notes) - 10} 条复核提示，请打开 Excel 核对。" if len(review_notes) > 10 else ""
        text += f"\n以下交易建议人工复核：\n{preview}{more}\n"
    else:
        text += "\n所有凭证均已通过规则或 AI 辅助匹配，借贷平衡。\n"

    return {
        "text": text,
        "file_name": file_doc.file_name,
        "file_url": file_doc.file_url,
        "transaction_count": len(transactions),
        "voucher_row_count": len(voucher_rows),
        "review_count": len(review_notes),
        "ai_candidate_count": ai_stats["candidate_count"],
        "ai_applied_count": ai_stats["applied_count"],
        "ai_invalid_count": ai_stats["invalid_count"],
    }
